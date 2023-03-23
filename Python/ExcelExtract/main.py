from turtle import pd
import pandas
import xlrd2
import xlwt
import openpyxl
import xlutils
import xlutils.copy


# 获取单个表格值 (2,1)表示获取第3行第2列单元格的值
# value = table.cell_value(2, 1)
# print("第3行2列值为", value)

# 获取表格行数


def write_TestDescriptionRx():
    row = 31
    listnum = 0
    column = 5
    # read
    xlsx = xlrd2.open_workbook('test.xlsx')
    table = xlsx.sheet_by_index(0)  # 通过索引查找：xlsx.sheet_by_index(3)
    nrows = table.nrows
    name_list = [str(table.cell_value(i, 1)) for i in range(0, nrows)]  # i代表行数 从0开始并且累加
    while '' in name_list:
        name_list.remove('')
    print(name_list)
    #以下代码暂时屏蔽为了测试读功能
    # new_name_list = pandas.Series(name_list) + ' Rx message test'
    # print(new_name_list)
    # # write
    # test = openpyxl.load_workbook('test1.xlsx')
    # sheet = test.worksheets[0]
    # while listnum < len(name_list):
    #     sheet.cell(row, column).value = new_name_list[listnum]
    #     test.save('test1.xlsx')
    #     listnum += 1
    #     row += 1


def write_TestDescriptionTx():
    row = 31
    listnum = 0
    column = 4
    # read
    xlsx = xlrd2.open_workbook('if.xlsx')
    table = xlsx.sheet_by_index(2)  # 通过索引查找：xlsx.sheet_by_index(3)
    nrows = table.nrows
    name_list = [str(table.cell_value(i, 13)) for i in range(1, nrows)]  # 获取第4列所有值（列表生成式）
    while '' in name_list:
        name_list.remove('')
    print(name_list)
    new_name_list = pandas.Series(name_list) + ' Tx message test'
    print(new_name_list)
    # write
    test = openpyxl.load_workbook('test1.xlsx')
    sheet = test.worksheets[0]
    while listnum < len(name_list):
        sheet.cell(row, column).value = new_name_list[listnum]
        test.save('test1.xlsx')
        listnum += 1
        row += 1

def write_Target():
    xlsx = xlrd2.open_workbook('if.xlsx')
    table = xlsx.sheet_by_index(2)
    list_Merge = table.merged_cells
    print(list_Merge)
    print(len(list_Merge))
    i = 0
    j = 0
    while i < len(list_Merge):
        tempRow = list_Merge[i][j]
        msgId = table.cell_value(list_Merge[i][1], list_Merge[i][2])
        print(msgId)
        print('Test the minimum, maximum and median values for all of the following variables')
        while tempRow < list_Merge[i][1]:
            var = table.cell_value(tempRow + 1, list_Merge[i][2] - 1)
            tempRow += 1
            print(var)
        i += 1


if __name__ == '__main__':
    write_TestDescriptionRx()
    # write_TestDescriptionRx()
